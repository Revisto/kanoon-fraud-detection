import openpyxl
import json
from pprint import pprint

class FetchData:
    def __init__(self, exam_file_path="Exams.xlsx", face_to_face_students_file_path="FTFStudents.xlsx", online_students_file_path="OStudents.xlsx"):
        self.exams_workbook = openpyxl.load_workbook(exam_file_path)
        self.face_to_face_students_answers_workbook = openpyxl.load_workbook(face_to_face_students_file_path)
        self.online_students_answers_workbook = openpyxl.load_workbook(online_students_file_path)

    def read_exams_answersheets(self):
        sheet = self.exams_workbook.worksheets[0]
        exams_answer_sheet = dict()
        for row in sheet:
            exam_name = row[0].value
            exam_answer_sheet = row[1].value
            exams_answer_sheet[exam_name] = json.loads(exam_answer_sheet)
        return exams_answer_sheet


    def read_face_to_face_students_answers(self):
        sheet = self.face_to_face_students_answers_workbook.worksheets[0]
        students_answers_to_exams = dict()
        for row in sheet:
            exam_name = row[0].value
            if exam_name not in students_answers_to_exams:
                students_answers_to_exams[exam_name] = list()
            
            student_answers = {
                "Exam": row[0].value,
                "Name": row[1].value,
                "Sheet": json.loads(row[2].value),
                "Date": row[3].value,
                "Duration": row[4].value
            }
            students_answers_to_exams[exam_name].append(student_answers)
        return students_answers_to_exams

    def read_online_students_answers(self):
        sheet = self.online_students_answers_workbook.worksheets[0]
        students_answers_to_exams = dict()
        for row in sheet:
            exam_name = row[0].value
            if exam_name not in students_answers_to_exams:
                students_answers_to_exams[exam_name] = list()
            
            student_answers = {
                "Exam": row[0].value,
                "Name": row[1].value,
                "Sheet": json.loads(row[2].value),
                "Date": row[3].value,
                "Durations": row[4].value,
                "IPs": row[5].value
            }
            students_answers_to_exams[exam_name].append(student_answers)
        return students_answers_to_exams




class FraudDetection:

    def __init__(self):
        self.suspects = dict()

    def add_red_flag(self, suspects, flag):
        for suspect in suspects:
            if suspect in self.suspects:
                self.suspects[suspect].append(flag)
            else:
                self.suspects[suspect] = list()
                self.suspects[suspect].append(flag)


    def find_similarity(self, sheet1, sheet2):
        suspect_rate = 0
        identical_answers_string = str()
        for answer_index in range(len(sheet1)):
            if sheet1[answer_index] == sheet2[answer_index]:
                identical_answers_string += "1"
            else:
                identical_answers_string += "0"
        identical_answers_ranges = identical_answers_string.split("0")
        identical_answers_ranges = [x for x in identical_answers_ranges if x]
        for identical_answer_range in identical_answers_ranges:
            suspect_rate += len(identical_answer_range) ** 2

        suspect_rate = suspect_rate ** .5
        suspect_rate = suspect_rate / len(sheet1)
        return suspect_rate

    def find_similar_sheets(self, exam_dict):
        cheaters = set()
        for suspect_index in range(len(exam_dict)):
            suspect = exam_dict[suspect_index]

            for second_suspect_index in range(suspect_index + 1, len(exam_dict)):
                second_suspect = exam_dict[second_suspect_index]
                similarity = self.find_similarity(suspect["Sheet"], second_suspect["Sheet"])
                if similarity > 00.08:
                    cheaters.add(suspect["Name"])
                    cheaters.add(second_suspect["Name"])
        self.add_red_flag(cheaters, "SIMILAR_ANSWER_SHEETS")
        return cheaters
    
    def find_same_ip_for_multiple_users(self, exam_dict):
        cheaters = set()
        all_ips = list()
        for student in exam_dict:
            student_ips = student["IPs"]
            student_ips = json.loads(student_ips)
            student_ips = list(set(student_ips))
            all_ips += student_ips
        
        for student in exam_dict:
            student_ips = student["IPs"]
            student_ips = json.loads(student_ips)
            student_ips = list(set(student_ips))
            for student_ip in student_ips:
                if all_ips.count(student_ip) > 1:
                    cheaters.add(student["Name"])

        self.add_red_flag(cheaters, "SAME IP")
        return cheaters

face_to_face_exams = FetchData().read_face_to_face_students_answers()
online_exams = FetchData().read_online_students_answers()
fraud_detection = FraudDetection()

#for exam_name in face_to_face_exams:
#    print(fraud_detection.find_similar_sheets(face_to_face_exams[exam_name]))
#    print(fraud_detection.suspects)

for exam_name in online_exams:
    print(fraud_detection.find_same_ip_for_multiple_users(online_exams[exam_name]))
    print(fraud_detection.suspects)